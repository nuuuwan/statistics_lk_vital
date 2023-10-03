from dataclasses import dataclass
from functools import cached_property

from openpyxl import Workbook
from utils import Log

from statistics_lk_vital.core.AgeRange import AgeRange
from statistics_lk_vital.core.Description import Description
from statistics_lk_vital.core.Gender import Gender

log = Log('MortalityStatistic')


@dataclass
class MortalityStatistic:
    description_raw: str
    year: int
    gender_raw: str
    age_raw: str
    deaths: int

    @cached_property
    def gender(self) -> str:
        return Gender.parse(self.gender_raw)

    @cached_property
    def age_range(self) -> str:
        return AgeRange.parse(self.age_raw)

    @staticmethod
    def get_idx_gda(statistics: list['MortalityStatistic']) -> dict:
        idx_gda = {}
        for statistic in statistics:
            gender = statistic.gender
            if gender is None:
                continue
            gender = str(gender)

            age_range = statistic.age_range
            if age_range is None:
                continue
            age_range = str(age_range)

            description = statistic.description_raw

            # idx_gda
            idx_gda[gender] = idx_gda.get(gender, {})
            idx_gda[gender][description] = idx_gda[gender].get(
                description, {}
            )
            idx_gda[gender][description][age_range] = idx_gda[gender][
                description
            ].get(age_range, 0)
            idx_gda[gender][description][age_range] += statistic.deaths
        return idx_gda

    @staticmethod
    def write_sheet(wb, gender, idx_da):
        ws = wb.create_sheet(gender)
        n_descriptions = len(idx_da.keys())

        ws.cell(1, 1).value = 'Row Code'
        ws.cell(1, 2).value = 'Long'
        ws.cell(1, 3).value = 'Short'
        offset_i_col_data = 4
        i_row = -1

        for description in idx_da.keys():
            _description = Description(description)
            if not _description.is_top_level:
                continue
            i_row += 1

            ws.cell(i_row + 2, 1).value = _description.row_code
            ws.cell(i_row + 2, 2).value = _description.details
            ws.cell(i_row + 2, 3).value = _description.simple

            for i_col, age_range in enumerate(idx_da[description].keys()):
                if i_row == 0:
                    ws.cell(1, i_col + offset_i_col_data).value = age_range

                deaths = idx_da[description][age_range]
                ws.cell(i_row + 2, i_col + offset_i_col_data).value = deaths

        log.debug(f'Wrote {n_descriptions} descriptions for {gender}')

    @staticmethod
    def write(statistics: list['MortalityStatistic'], file_path: str):
        assert file_path.endswith('.xlsx')

        idx_gda = MortalityStatistic.get_idx_gda(statistics)

        wb = Workbook()
        for gender, idx_da in idx_gda.items():
            MortalityStatistic.write_sheet(wb, gender, idx_da)

        wb.remove(wb['Sheet'])

        wb.save(file_path)
        log.info(f'Wrote {len(statistics)} statistics to {file_path}')
