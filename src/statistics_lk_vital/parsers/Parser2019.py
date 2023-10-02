from openpyxl import load_workbook
from utils import Log

from statistics_lk_vital.core.MortalityStatistic import MortalityStatistic

log = Log('Parser2019')


def parse_int(x):
    try:
        return int(x)
    except Exception:
        return 0


def parse_age(x: str) -> tuple[int, int]:
    age_min, age_max = None, None

    for over_token in ['and over', '+']:
        if over_token in x:
            age_min = x.replace(over_token, '').strip()

    if '-' in x:
        age_min, age_max = x.replace(' ', '').split('-')
    return age_min, age_max


def parse_death_code(x: str) -> tuple[str, str]:
    if '-' in x:
        death_code_min, death_code_max = x.split('-')
    else:
        death_code_min = death_code_max = x
    return death_code_min, death_code_max


def parse_col_to_age(cells):
    col_to_age = []
    non_none_cell = None
    for cell in cells:
        if cell is not None:
            non_none_cell = cell
        col_to_age.append(non_none_cell)
    return col_to_age


def parse_col_to_gender(cells):
    col_to_gender = []
    non_none_cell = None
    for cell in cells:
        if cell is not None:
            non_none_cell = (
                cell.replace('Total', '').replace('years', '').strip()
            )
        col_to_gender.append(non_none_cell)
    return col_to_gender


class Parser2019:
    def __init__(self, year: str, file_path: str):
        self.year = year
        self.file_path = file_path

    def parse(self) -> list[MortalityStatistic]:
        workbook = load_workbook(filename=self.file_path)
        sheet = workbook.active
        statistics = []
        i_row = 0
        col_to_age = None
        col_to_gender = None
        for i_row, row in enumerate(sheet.iter_rows(values_only=True)):
            if i_row < 3:
                continue
            cells = list(row)

            if i_row == 3:  # Age:
                col_to_age = parse_col_to_age(cells)
                continue

            if i_row == 4:  # Gender
                col_to_gender = parse_col_to_gender(cells)
                continue

            title_row = cells[0]
            if title_row is None:
                continue

            tokens = title_row.split(' ')

            death_code_min, death_code_max = parse_death_code(tokens[-1])
            death_description = ' '.join(tokens[1:-1])

            for i_cell, cell in enumerate(cells):
                if i_cell == 0:
                    continue

                age_min, age_max = parse_age(col_to_age[i_cell])

                gender = col_to_gender[i_cell].strip()
                deaths = parse_int(cell)
                statistic = MortalityStatistic(
                    death_description=death_description,
                    death_code_min=death_code_min,
                    death_code_max=death_code_max,
                    year=self.year,
                    gender=gender,
                    age_min=age_min,
                    age_max=age_max,
                    deaths=deaths,
                )
                statistics.append(statistic)
                log.debug(str(statistic))
        return statistics


if __name__ == '__main__':
    parser = Parser2019(
        year='2019', file_path='data/cause-of-death-2019.xlsx'
    )
    parser.parse()
